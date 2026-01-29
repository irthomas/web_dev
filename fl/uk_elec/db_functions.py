# -*- coding: utf-8 -*-
"""
Created on Fri Mar 19 10:33:11 2021

@author: iant

DB FUNCTIONS
"""



import sqlite3
import re
import os
# import sys
# from datetime import datetime

from config import DB_PATH, APP_PATH
import numpy as np

import matplotlib.pyplot as plt
from openpyxl import load_workbook
   
    
def connect_db():
    con = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
    return con

def close_db(con):
    con.close()


def regexp(exp, regex_str):
    regex = re.compile(exp)
    return regex.search(regex_str) is not None




def search_db(table_name, search_params):
    """search filename field for regex"""
    # con = connect_db()
    # con.create_function("REGEXP", 2, regexp)
    # cur = con.cursor()
    # cur.execute('SELECT * FROM {} WHERE filename REGEXP ?'.format(table_name), [regex_str])

    print(search_params)
    con = connect_db()
    cur = con.cursor()
    cur.execute('SELECT * FROM {} WHERE utc_start_time BETWEEN ? AND ? AND mtp_number BETWEEN ? AND ?'.format(table_name), \
        [
            search_params["utc_start_time_s"], 
            search_params["utc_start_time_e"],
            search_params["mtp_s"],
            search_params["mtp_e"],
            
        ])


    rows = cur.fetchall()
    close_db(con)
    return rows


WB_INFO = {
    2017:{
        "data":{
            "header_row":3, 
            "header_cols":range(1, 62), 
            "data_rows":range(4,654), 
            "sheet":"Administrative data", 
            "path":os.path.join(APP_PATH, "db", "2017-UKPGE-Electoral-Data.xlsx"),
            "indices":{}
        }, 
        "results":{
            "header_row":2, 
            "header_cols":range(1, 9), 
            "data_rows":range(3,3307), 
            "sheet":"Results", 
            "path":os.path.join(APP_PATH, "db", "2017-UKPGE-Electoral-Data.xlsx"),
            "indices":{"constituency":2, "party":6, "votes":7}
        }
    },
    2019:{
        "results":{
            "header_row":1, 
            "header_cols":range(1, 18), 
            "data_rows":range(2,3322), 
            "sheet":"Sheet1", 
            "path":os.path.join(APP_PATH, "db", "HoC-GE2019-results-by-candidate-xlsx.xlsx"),
            "indices":{"constituency":2, "party":7, "votes":14}
        }
    }
}


def get_xlsx_data(year, key):
    
    wb = load_workbook(WB_INFO[year][key]["path"], data_only=True)
    # sheets = wb.sheetnames
    sheet = wb[WB_INFO[year][key]["sheet"]]
    
    headers = []
    for j in WB_INFO[year][key]["header_cols"]:
        headers.append(sheet.cell(WB_INFO[year][key]["header_row"], j).value)
        
    data = []
    for i in WB_INFO[year][key]["data_rows"]:
        data_row = []
        for j in WB_INFO[year][key]["header_cols"]:
            data_row.append(sheet.cell(i, j).value)
        
        data.append(data_row)
    
    return headers, data, WB_INFO[year][key]["indices"]


#admin data
# a_headers, a_data = get_xlsx_data(XLSX_PATHS[0], "Administrative data")

#results_data
r_headers, r_data, r_indices = get_xlsx_data(2017, "results")


#sort by constituency
constituencies = sorted(list(set([i[r_indices["constituency"]] for i in r_data])))

c_dict = {c:{"party":[], "votes":[]} for c in constituencies}

for row in r_data:
    c_dict[row[r_indices["constituency"]]]["party"].append(row[r_indices["party"]])
    c_dict[row[r_indices["constituency"]]]["votes"].append(row[r_indices["votes"]])
    
    

for c, c_data in c_dict.items():
    votes = np.array(c_data["votes"])
    vote_indices = np.argsort(votes)
    c_data["majority"] = votes[vote_indices[-1]] - votes[vote_indices[-2]]
    c_data["winner"] = c_data["party"][vote_indices[-1]]
    

majorities = np.array([value["majority"] for key, value in c_dict.items()])
winners = np.array([value["winner"] for key, value in c_dict.items()])
majority_indices = np.argsort(majorities)


colour_dict = {"Labour":"r", "Liberal Democrats":"y", "SNP":"orange", "DUP":"g", "Conservative":"b"}
colours = []
for winner in winners:
    if winner in colour_dict.keys():
        colours.append(colour_dict[winner])
    else:
        colours.append("k")
colours = np.array(colours)


#get indices for L and C
l_indices = [m for i, (m, w) in enumerate(zip(majority_indices, winners[majority_indices])) if w == "Labour"]
c_indices = [m for i, (m, w) in enumerate(zip(majority_indices, winners[majority_indices])) if w == "Conservative"]

plt.figure()
plt.scatter(range(len(colours)), majorities[majority_indices], color=colours[majority_indices])
plt.grid()

plt.figure()
plt.scatter(range(len(l_indices[::-1])), majorities[l_indices[::-1]], color="r")
plt.scatter(range(len(l_indices), len(l_indices)+len(c_indices)), -1*majorities[c_indices], color="b")
plt.grid()

#L swing needs = 28 seats
swing = int(np.ceil((len(c_indices) - len(l_indices))/2.))

#total L votes needed (assuming C in 2nd place)
sum(majorities[l_indices][0:swing]) # 23420 

#23420 vote swing = 28 extra C seats = 317 + 28 = 345 seats total  (assuming L in 2nd place)
sum(majorities[c_indices][0:28]) #= 23099



#results_data
r_headers, r_data, r_indices = get_xlsx_data(2019, "results")


#sort by constituency
constituencies = sorted(list(set([i[r_indices["constituency"]] for i in r_data])))

c_dict = {c:{"party":[], "votes":[]} for c in constituencies}

for row in r_data:
    c_dict[row[r_indices["constituency"]]]["party"].append(row[r_indices["party"]])
    c_dict[row[r_indices["constituency"]]]["votes"].append(row[r_indices["votes"]])
    
    

for c, c_data in c_dict.items():
    votes = np.array(c_data["votes"])
    vote_indices = np.argsort(votes)
    c_data["majority"] = votes[vote_indices[-1]] - votes[vote_indices[-2]]
    c_data["winner"] = c_data["party"][vote_indices[-1]]
    

majorities = np.array([value["majority"] for key, value in c_dict.items()])
winners = np.array([value["winner"] for key, value in c_dict.items()])
majority_indices = np.argsort(majorities)


colour_dict = {"Labour":"r", "Liberal Democrats":"y", "SNP":"orange", "DUP":"g", "Conservative":"b"}
colours = []
for winner in winners:
    if winner in colour_dict.keys():
        colours.append(colour_dict[winner])
    else:
        colours.append("k")
colours = np.array(colours)


#get indices for L and C
l_indices = [m for i, (m, w) in enumerate(zip(majority_indices, winners[majority_indices])) if w == "Labour"]
c_indices = [m for i, (m, w) in enumerate(zip(majority_indices, winners[majority_indices])) if w == "Conservative"]

plt.figure()
plt.scatter(range(len(colours)), majorities[majority_indices], color=colours[majority_indices])
plt.grid()

plt.figure()
plt.scatter(range(len(l_indices[::-1])), majorities[l_indices[::-1]], color="r")
plt.scatter(range(len(l_indices), len(l_indices)+len(c_indices)), -1*majorities[c_indices], color="b")
plt.grid()